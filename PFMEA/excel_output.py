import io
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)

# 列定義：(列記号, ヘッダー表示名, DBカラム名 or None)
COLUMNS = [
    ("B", "No.",              "id"),
    ("C", "工程",              None),
    ("D", "工程の役割",         "process"),
    ("E", "故障モード",         "failure_mode"),
    ("F", "故障の影響",         "effect"),
    ("G", "厳しさ（S）",        "severity"),
    ("H", "重要区分",           None),
    ("I", "故障原因／メカニズム", "cause"),
    ("J", "発生頻度（O）",       "occurrence"),
    ("K", "発生予防",           "current_control_prevention"),
    ("L", "故障の検出",         "current_control_detection"),
    ("M", "検出度（D）",        "detection"),
    ("N", "RPN",              "rpn"),
]

# スタイル定義
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

THIN_SIDE    = Side(style="thin", color="AAAAAA")
THIN_BORDER  = Border(
    top=THIN_SIDE, bottom=THIN_SIDE,
    left=THIN_SIDE, right=THIN_SIDE
)

# 列幅定義（文字数相当）
COL_WIDTHS = {
    "B": 6,
    "C": 10,
    "D": 20,
    "E": 25,
    "F": 30,
    "G": 8,
    "H": 10,
    "I": 30,
    "J": 8,
    "K": 30,
    "L": 30,
    "M": 8,
    "N": 8,
}

def build_excel(
    records: list[dict],
    industry: str = "",
    product: str = ""
) -> bytes:
    """
    レコードリストからExcelファイルを生成し、bytesで返す
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # シート名
    sheet_name = f"PFMEA_{industry}_{product}"[:31]  # Excel上限31文字
    ws.title = sheet_name

    # 行1：タイトル行
    ws.merge_cells("B1:N1")
    title_cell = ws["B1"]
    title_cell.value = f"PFMEA　{industry}　{product}　出力日：{datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font  = Font(name="Arial", bold=True, size=12, color="1F4E79")
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 20

    # 行2：ヘッダー行
    for col_letter, header, _ in COLUMNS:
        cell = ws[f"{col_letter}2"]
        cell.value     = header
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER
    ws.row_dimensions[2].height = 20

    # 行3以降：データ行
    for row_idx, record in enumerate(records, start=3):
        for col_letter, _, db_key in COLUMNS:
            cell = ws[f"{col_letter}{row_idx}"]
            if db_key is None:
                cell.value = ""
            else:
                cell.value = record.get(db_key, "")

            cell.font   = DATA_FONT
            cell.border = THIN_BORDER

            # 数値列はセンタリング
            if col_letter in ("B", "G", "H", "J", "M", "N"):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = WRAP_ALIGN

        # 行の高さを自動調整（折り返しテキスト対応）
        ws.row_dimensions[row_idx].height = 45

    # 列幅設定
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ウィンドウ枠の固定（ヘッダー行を固定）
    ws.freeze_panes = "B3"

    # bytesとして返す
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def make_filename(industry: str, product: str) -> str:
    """
    ダウンロード用ファイル名を生成する
    """
    date_str = datetime.now().strftime("%Y%m%d")
    return f"PFMEA_{industry}_{product}_{date_str}.xlsx"
